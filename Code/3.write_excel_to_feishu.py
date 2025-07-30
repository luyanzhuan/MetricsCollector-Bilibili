#!/usr/bin/env python3
# -*- coding: utf-8 -*-
'''
Author       : luyz
Date         : 2025-07-26 21:39:40
LastEditors  : luyz
LastEditTime : 2025-07-30 00:01:40
Description  : 撰写 Excel 数据到飞书表格
Copyright (c) 2025 by LuYanzhuan lyanzhuan@gmail.com, All Rights Reserved.
'''

import requests
import argparse
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# ====================== 函数定义部分 ======================

def get_tenant_access_token(app_id, app_secret):
    """
    获取 tenant_access_token，用于访问飞书开放平台 API。
    """
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal/"
    payload = {
        "app_id": app_id,
        "app_secret": app_secret
    }
    try:
        response = requests.post(url, json=payload)
        data = response.json()
    except Exception as e:
        raise RuntimeError(f"请求 tenant_access_token 失败: {e}")
    
    if response.status_code != 200 or "tenant_access_token" not in data:
        raise RuntimeError(f"获取 tenant_access_token 出错: {response.status_code}, 响应: {response.text}")
    
    return data["tenant_access_token"]

def read_excel_data(file_path, sheet_name):
    """
    从本地 Excel 文件读取数据，返回二维列表。
    将 datetime 类型转为字符串，防止 JSON 报错。
    """
    wb = load_workbook(file_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel文件中不存在名为'{sheet_name}'的工作表")
    
    ws = wb[sheet_name]
    data = []

    def convert_cell(cell):
        if isinstance(cell, datetime):
            return cell.strftime("%Y-%m-%d %H:%M:%S")
        elif cell is None:
            return ""
        else:
            return cell

    for row in ws.iter_rows(values_only=True):
        data.append([convert_cell(cell) for cell in row])

    wb.close()
    return data

def write_to_feishu_sheet(token, spreadsheet_token, sheet_id, start_cell, values):
    """
    将数据写入飞书 Sheets 表格。
    """
    col_letters = ''.join(filter(str.isalpha, start_cell))
    row_start = int(''.join(filter(str.isdigit, start_cell)))
    col_start_index = column_index_from_string(col_letters) if col_letters else 1

    num_rows = len(values)
    num_cols = len(values[0]) if num_rows > 0 else 0
    end_row = row_start + num_rows - 1
    end_col_index = col_start_index + num_cols - 1
    end_col_letters = get_column_letter(end_col_index)
    end_cell = f"{end_col_letters}{end_row}"
    range_str = f"{sheet_id}!{col_letters}{row_start}:{end_cell}"

    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}"
    }
    payload = {
        "valueRanges": [
            {
                "range": range_str,
                "values": values
            }
        ]
    }

    response = requests.post(url, json=payload, headers=headers)
    result = response.json()
    if result.get("code", -1) != 0:
        raise RuntimeError(f"写入飞书表格失败: 错误码 {result.get('code')}, 信息: {result.get('msg')}")
    return result

# ====================== 主程序入口 ======================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="将 Excel 内容写入飞书表格")

    # 命令行参数配置
    parser.add_argument("--excel_path", required=True, help="Excel 文件路径")
    parser.add_argument("--sheet_name", required=True, help="Excel 工作表名称")
    parser.add_argument("--app_id", required=True, help="飞书开放平台 App ID")
    parser.add_argument("--app_secret", required=True, help="飞书开放平台 App Secret")
    parser.add_argument("--spreadsheet_token", required=True, help="飞书表格 Spreadsheet Token")
    parser.add_argument("--sheet_id", required=True, help="飞书表格 Sheet ID")
    parser.add_argument("--start_cell", default="A1", help="写入起始单元格（默认为 A1）")

    args = parser.parse_args()

    try:
        # 读取数据
        data_values = read_excel_data(args.excel_path, args.sheet_name)
        print(f"读取 Excel 成功，共 {len(data_values)} 行数据")

        # 获取 Token
        token = get_tenant_access_token(args.app_id, args.app_secret)
        print("获取 tenant_access_token 成功")

        # 写入飞书表格
        result = write_to_feishu_sheet(token, args.spreadsheet_token, args.sheet_id, args.start_cell, data_values)
        print("✅ 成功写入飞书表格:", result.get("msg", ""))
    except Exception as e:
        print("❌ 脚本执行失败:", e)
