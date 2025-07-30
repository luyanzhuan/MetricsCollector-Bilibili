import requests
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

# ========= 配置部分（用户需根据自身情况修改） =========
APP_ID = "cli_a80266874478100b"
APP_SECRET = "hQEQkgXSTjGTOQNDocCkRNM0GSjoDqGH"

# 飞书表格的标识和位置参数
SPREADSHEET_TOKEN = "DIxIsYSXXhJAxDtgV20coEbWnse"
SHEET_ID = "079e96"
START_CELL = "A1"  # 写入的起始单元格位置，例如"A1"

# 本地 Excel 文件路径和Sheet名称
EXCEL_FILE_PATH = "/data2/luyz/project/CustomizedAnalysis/CustomizedAnalysis.12.20250726_YASbilibili/Test/video_stats_output/today.xlsx"         # 替换为你的Excel文件路径
EXCEL_SHEET_NAME = "Sheet1"           # 替换为要读取的工作表名称

# ========= 函数定义部分 =========

def get_tenant_access_token(app_id, app_secret):
    """
    调用飞书开放平台获取 tenant_access_token。
    需要应用的 App ID 和 App Secret。返回 tenant_access_token 字符串。
    """
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal/"
    # 构造请求数据
    payload = {
        "app_id": app_id,
        "app_secret": app_secret
    }
    try:
        response = requests.post(url, json=payload)
        data = response.json()
    except Exception as e:
        raise RuntimeError(f"请求 tenant_access_token 失败: {e}")
    # 检查响应
    if response.status_code != 200 or "tenant_access_token" not in data:
        raise RuntimeError(f"获取 tenant_access_token 出错: {response.status_code}, 响应: {response.text}")
    token = data["tenant_access_token"]
    return token

def read_excel_data(file_path, sheet_name):
    """
    读取本地 Excel 文件中特定工作表的数据。
    返回一个二维列表，每个子列表代表Excel中的一行。
    同时将 datetime 类型转为字符串，避免 JSON 序列化报错。
    """
    wb = load_workbook(file_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Excel文件中不存在名为'{sheet_name}'的工作表")
    ws = wb[sheet_name]
    data = []

    def convert_cell(cell):
        if isinstance(cell, datetime):
            return cell.strftime("%Y-%m-%d %H:%M:%S")  # 或 cell.isoformat()
        elif cell is None:
            return ""
        else:
            return cell

    for row in ws.iter_rows(values_only=True):
        data.append([convert_cell(cell) for cell in row])

    wb.close()
    return data

def write_to_feishu_sheet(token, spreadsheet_token, sheet_id, start_cell, values):
    """
    调用飞书 Sheets API，将指定的数据写入飞书表格。
    token 是调用API所需的 tenant_access_token。
    spreadsheet_token 是飞书表格文档的 token 标识。
    sheet_id 是目标工作表的 ID。
    start_cell 是起始单元格位置（如"A1"）。
    values 是二维列表，表示要写入的矩阵数据。
    """
    # 计算数据范围的结束单元格坐标
    # 解析起始单元格，例如 "B2" -> col_start = 2, row_start = 2
    col_letters = ''.join(filter(str.isalpha, start_cell))
    row_start = int(''.join(filter(str.isdigit, start_cell)))
    col_start_index = column_index_from_string(col_letters) if col_letters else 1
    # 数据维度
    num_rows = len(values)
    num_cols = len(values[0]) if num_rows > 0 else 0
    # 计算结束单元格的行号和列字母
    end_row = row_start + num_rows - 1
    end_col_index = col_start_index + num_cols - 1
    end_col_letters = get_column_letter(end_col_index)
    end_cell = f"{end_col_letters}{end_row}"
    # 构造 range 字符串，例如 "402cb1!A1:D10"
    range_str = f"{sheet_id}!{col_letters}{row_start}:{end_cell}"
    
    # 构造请求URL和数据
    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values_batch_update"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}"
    }
    payload = {
        "valueRanges": [
            {
                "range": range_str,
                "values": values
            }
        ]
    }
    # 发送请求
    response = requests.post(url, json=payload, headers=headers)
    result = response.json()
    print(result)
    if result.get("code", -1) != 0:
        # API返回错误，抛出异常并打印错误信息
        raise RuntimeError(f"写入飞书表格失败: 错误码 {result.get('code')}, 信息: {result.get('msg')}")
    return result

# ========= 主程序执行部分 =========

if __name__ == "__main__":
    try:
        # 1. 读取Excel数据
        data_values = read_excel_data(EXCEL_FILE_PATH, EXCEL_SHEET_NAME)
        print(f"从Excel读取{len(data_values)}行数据，将写入飞书表格...")
        # 2. 获取tenant_access_token
        tenant_token = get_tenant_access_token(APP_ID, APP_SECRET)
        print("成功获取 tenant_access_token")
        # 3. 写入数据到飞书表格
        result = write_to_feishu_sheet(tenant_token, SPREADSHEET_TOKEN, SHEET_ID, START_CELL, data_values)
        print("数据已成功写入飞书表格！返回信息:", result.get("msg", ""))
    except Exception as err:
        print("脚本执行过程中发生错误:", err)
